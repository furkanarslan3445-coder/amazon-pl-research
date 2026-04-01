import json
import glob
import os
from flask import Flask, jsonify, send_file
from openpyxl import load_workbook
from config import CONFIG

app = Flask(__name__)


def load_json(path, default):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return default


@app.route("/")
def index():
    return send_file("dashboard.html")


@app.route("/api/stats")
def stats():
    pool = load_json(CONFIG["keywords_pool"], [])
    done = load_json(CONFIG["keywords_done"], [])
    opportunities = get_opportunities()
    return jsonify({
        "total": len(pool),
        "scanned": len(done),
        "remaining": max(0, len(pool) - len(done)),
        "opportunities": len(opportunities),
    })


@app.route("/api/recent")
def recent():
    done = load_json(CONFIG["keywords_done"], [])
    return jsonify(done[-50:][::-1])  # son 50, en yeniden


@app.route("/api/opportunities")
def opportunities():
    return jsonify(get_opportunities())


def get_opportunities():
    files = glob.glob(os.path.join(CONFIG["output_dir"], "*.xlsx"))
    if not files:
        return []
    results = []
    for f in files:
        try:
            wb = load_workbook(f)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    results.append({
                        "keyword": row[0],
                        "product_count": row[1],
                        "avg_price": row[2],
                        "opportunity_count": row[3],
                        "ratio": row[4],
                    })
        except Exception:
            pass
    return results


if __name__ == "__main__":
    app.run(port=5050, debug=False)
