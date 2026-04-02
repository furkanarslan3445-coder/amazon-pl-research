import os
import subprocess
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from config import CONFIG


HEADERS = [
    "Anahtar Kelime", "Ürün Sayısı", "Ort. Fiyat ($)",
    "Uygun Ürün Sayısı", "Oran (%)", "Durum"
]

GREEN  = "FF00AA44"
RED    = "FFCC0000"
YELLOW = "FFFFD700"
GRAY   = "FF2D2D2D"


def _get_or_create_wb() -> tuple:
    path = CONFIG["output_file"]
    os.makedirs(CONFIG["output_dir"], exist_ok=True)

    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Niche Firsatlari"
        _write_header(ws)

    return wb, ws, path


def _write_header(ws):
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFFFF")
        cell.fill      = PatternFill("solid", fgColor=GRAY)
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 35
    for col in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 18


def _save_json(ws):
    import json
    import glob
    rows = []
    seen = set()
    # Tüm Excel dosyalarını oku
    for f in glob.glob(os.path.join(CONFIG["output_dir"], "*.xlsx")):
        try:
            wb = load_workbook(f)
            for row in wb.active.iter_rows(min_row=2, values_only=True):
                if row[0] and row[0] not in seen:
                    seen.add(row[0])
                    rows.append({
                        "keyword": row[0],
                        "product_count": row[1],
                        "avg_price": row[2],
                        "opportunity_count": row[3],
                        "ratio": row[4],
                    })
        except Exception:
            pass
    json_path = os.path.join(CONFIG["output_dir"], "opportunities.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)


def _git_push(path: str):
    try:
        json_path = os.path.join(CONFIG["output_dir"], "opportunities.json")
        subprocess.run(["git", "add", path, json_path], check=True, capture_output=True)
        subprocess.run(["git", "commit", "-m", "yeni fırsat eklendi"], check=True, capture_output=True)
        subprocess.run(["git", "push"], check=True, capture_output=True)
    except Exception:
        pass


BLACKLIST = ["tablet", "ipad", "laptop", "phone", "smartphone", "iphone", "macbook",
             "minoxidil", "samsung", "3 days", "3 day"]


def save_opportunity(result: dict):
    """Fırsat bulunan keyword'ü Excel'e ekle."""
    kw = result["keyword"].lower()
    if any(b in kw for b in BLACKLIST):
        return
    wb, ws, path = _get_or_create_wb()

    status = "✅ GİRİLEBİLİR" if result["is_opportunity"] else "❌"
    row = [
        result["keyword"],
        result["product_count"],
        f"${result['avg_price']:.2f}",
        result["opportunity_count"],
        f"{result['opportunity_ratio']}%",
        status,
    ]
    ws.append(row)

    # Son satırı renklendir
    last_row = ws.max_row
    color = GREEN if result["is_opportunity"] else RED
    for cell in ws[last_row]:
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(color="FFFFFFFF")

    wb.save(path)
    _save_json(ws)
    _git_push(path)
